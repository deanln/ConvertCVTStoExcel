from openpyxl import load_workbook
import shutil, os
import datetime
import extractCanMessages


def convertToExcel(pathOfTemplate, dataList):
    """
    Converts a list of logEntry objects into an excel sheet
    """

    homepath = os.getenv('USERPROFILE')

    fileName = createCopyOfTemplate(pathOfTemplate)
    workbook = openWorkbook(fileName)
    populateWorkbook(workbook, dataList)
    workbook.save(homepath + '\\Desktop\\' + fileName)
    #workbook.save(fileName)

def createCopyOfTemplate(pathOfTemplate):
    """
    Creates copy of file and returns name of the newly created file
    """
    fileName = 'CVTS NVM ' + datetime.datetime.now().strftime("%I.%M%p - %m.%d.%Y") + ".xlsx"
    shutil.copy(pathOfTemplate, fileName)
    return fileName

def openWorkbook(fileName):
    """
    Opens existing workbook and returns
    """
    workbook = load_workbook(fileName)
    return workbook

def populateWorkbook(workbook, dataList):
    """
    Populates the excel sheet with a list of logEntry objects
    """
    startRow = 4
    worksheet = workbook["Data Parser"]
    for entry in dataList:
        worksheet['B' + str(startRow)] = entry.getFaultSymbol()[1]
        worksheet['C' + str(startRow)] = entry.getMsg3()[0]
        worksheet['D' + str(startRow)] = entry.getMsg3()[1]
        worksheet['C' + str(startRow + 1)] = entry.getMsg4()[0]
        worksheet['D' + str(startRow + 1)] = entry.getMsg4()[1]
        worksheet['C' + str(startRow + 2)] = entry.getMsg7()[0]
        worksheet['D' + str(startRow + 2)] = entry.getMsg7()[1]
        startRow += 3
    

if __name__ == '__main__':
    fileName = extractCanMessages.createDataList(r"[inputfilehere]")
    convertToExcel("dataTemplate.xlsx",fileName)
